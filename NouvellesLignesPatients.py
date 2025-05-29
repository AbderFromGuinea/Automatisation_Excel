#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script pour extraire les nouvelles lignes manquantes entre un fichier Excel principal
et plusieurs fichiers Excel sources.

Ce script effectue les opérations suivantes :
1.  Identifie tous les fichiers Excel (.xlsx) dans le répertoire courant
    qui serviront de fichiers sources.
2.  Pour chaque fichier source :
    a.  Charge le fichier principal ("Recette.xlsx") et le fichier source actuel.
    b.  Recherche dynamiquement les colonnes 'Date' et 'Hôpital' (ou variantes)
        dans les en-têtes des deux fichiers.
    c.  Identifie un hôpital de référence à partir du fichier source (la première
        occurrence trouvée).
    d.  Récupère toutes les dates associées à cet hôpital dans le fichier source
        et dans le fichier principal.
    e.  Détermine les dates présentes dans le fichier source mais absentes du
        fichier principal pour cet hôpital.
    f.  Si des dates manquantes sont trouvées, il crée un nouveau fichier Excel.
    g.  Copie l'en-tête (du fichier principal par défaut) dans ce nouveau fichier.
    h.  Copie uniquement les lignes du fichier source qui correspondent à
        l'hôpital de référence et aux dates manquantes dans le nouveau fichier.
    i.  Enregistre le nouveau fichier avec un nom séquentiel (par exemple,
        new_lines_only1.xlsx, new_lines_only2.xlsx, etc.).
3.  Affiche le nombre de lignes ajoutées pour chaque fichier traité et des messages
    d'information ou d'erreur.
"""

from datetime import datetime
from openpyxl import load_workbook, Workbook
import glob # Pour trouver des fichiers correspondant à un motif (pattern matching)
import os   # Pour les opérations liées au système d'exploitation, comme la vérification de l'existence de fichiers

# --- Configuration des Fichiers ---
PRINCIPAL = "Ventes1.xlsx"  # Nom du fichier Excel principal de référence

# --- Fonctions Utilitaires ---

def find_column(ws, keywords, max_col=8):
    """
    Recherche une colonne dans la première ligne (en-tête) d'une feuille Excel
    en se basant sur une liste de mots-clés. La recherche est insensible à la casse.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): La feuille Excel où chercher.
        keywords (list): Une liste de chaînes de caractères (mots-clés) à rechercher.
        max_col (int, optional): Le nombre maximum de colonnes à scanner.
                                 Par défaut à 8.

    Returns:
        int: Le numéro de la colonne (1-indexed) si un mot-clé est trouvé.

    Raises:
        ValueError: Si aucun des mots-clés n'est trouvé dans les en-têtes
                    des colonnes spécifiées.
    """
    for col in range(1, max_col + 1):  # Parcourt les colonnes de 1 à max_col
        header_cell = ws.cell(row=1, column=col).value
        if header_cell:  # Vérifie si la cellule d'en-tête a une valeur
            # Vérifie si l'un des mots-clés (en minuscules) est présent dans la valeur d'en-tête (en minuscules)
            if any(keyword.lower() in str(header_cell).lower() for keyword in keywords):
                return col  # Retourne l'index de la colonne (1-based)
    # Si la boucle se termine sans trouver de colonne, lève une erreur
    raise ValueError(f"Colonne pour les mots-clés {keywords} non trouvée dans l'en-tête de {ws.parent.path}")


def excel_to_datetime(cell):
    """
    Convertit la valeur d'une cellule Excel en un objet datetime.
    Tente de parser les formats de date courants.

    Args:
        cell (openpyxl.cell.cell.Cell): La cellule Excel contenant la date.

    Returns:
        datetime.datetime or None: Un objet datetime si la conversion réussit,
                                   None sinon.
    """
    value = cell.value
    if isinstance(value, datetime):
        # La valeur est déjà un objet datetime (souvent le cas si openpyxl le gère bien)
        return value
    if isinstance(value, (int, float)) and cell.is_date:
        # Si c'est un nombre et qu'Excel le marque comme date (nombre de série Excel)
        # openpyxl le convertit généralement automatiquement en datetime,
        # mais cette vérification peut être utile.
        # Note: openpyxl > 2.4 gère cela nativement.
        # Pour les anciennes versions ou des cas spécifiques, une conversion pourrait être nécessaire ici.
        # from openpyxl.utils.datetime import from_excel
        # return from_excel(value)
        return value # En supposant qu'openpyxl a déjà fait la conversion.
    if isinstance(value, str):
        # Si la valeur est une chaîne, essaie de la parser
        try:
            return datetime.strptime(value, "%d/%m/%Y")
        except ValueError:
            try:
                # Essayer un autre format courant si le premier échoue
                return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    # Format ISO date
                    return datetime.strptime(value, "%Y-%m-%d")
                except ValueError:
                    # Si tous les formats échouent, retourne None
                    # Un message d'avertissement pourrait être ajouté ici si nécessaire
                    # print(f"Avertissement: Impossible de parser la date '{value}'")
                    return None
    return None # Retourne None si le type n'est pas géré ou la conversion échoue


def get_hospital_dates(ws, hosp_value):
    """
    Récupère et renvoie une liste triée de dates uniques pour un hôpital donné
    à partir d'une feuille Excel.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): La feuille Excel à analyser.
        hosp_value (str): Le nom de l'hôpital à rechercher.

    Returns:
        list: Une liste d'objets datetime, triée et sans doublons, pour l'hôpital spécifié.
              Retourne une liste vide si aucune date n'est trouvée ou si les colonnes
              requises ne sont pas trouvées.
    """
    try:
        # Trouve les colonnes 'Date' et 'Hôpital'
        date_col_idx = find_column(ws, ['date', 'DATE', 'Date'])
        hosp_col_idx = find_column(ws, ['HOPITAUX/ CMC', 'hospital', 'hopital', 'hôpital'])
    except ValueError as e:
        # Si les colonnes ne sont pas trouvées, affiche une erreur et retourne une liste vide
        print(f"Erreur lors de la recherche de colonnes dans {ws.parent.path}: {e}")
        return []

    dates = []
    # Parcourt les lignes de la feuille, en commençant par la deuxième ligne (après l'en-tête)
    for r in range(2, ws.max_row + 1):
        hospital_cell_value = ws.cell(row=r, column=hosp_col_idx).value
        # Vérifie si la valeur de la cellule de l'hôpital correspond à l'hôpital recherché
        if hospital_cell_value == hosp_value:
            date_cell = ws.cell(row=r, column=date_col_idx)
            datetime_obj = excel_to_datetime(date_cell) # Convertit la date
            if datetime_obj:
                dates.append(datetime_obj)
    # Retourne une liste de dates uniques (grâce à set()) et triées
    return sorted(list(set(dates)))


def find_missing_dates(source_dates, principal_dates):
    """
    Identifie les dates présentes dans la liste `source_dates` mais absentes
    de la liste `principal_dates`.

    Args:
        source_dates (list): Liste d'objets datetime provenant du fichier source.
        principal_dates (list): Liste d'objets datetime provenant du fichier principal.

    Returns:
        list: Une liste d'objets datetime qui sont dans `source_dates` mais pas
              dans `principal_dates`.
    """
    # Utilise une compréhension de liste pour trouver les dates manquantes
    return [date_obj for date_obj in source_dates if date_obj not in principal_dates]


# --- Fonction Principale de Traitement ---

def process_files(principal_filepath, source_filepath, output_filepath):
    """
    Traite un fichier source par rapport à un fichier principal pour en extraire
    les lignes avec des dates manquantes pour un hôpital spécifique.

    Args:
        principal_filepath (str): Chemin vers le fichier Excel principal.
        source_filepath (str): Chemin vers le fichier Excel source à traiter.
        output_filepath (str): Chemin où enregistrer le fichier Excel contenant
                               les nouvelles lignes.
    """
    print(f"\n--- Début du traitement pour le fichier source: {source_filepath} ---")
    try:
        # Charger les classeurs (workbooks) et sélectionner les feuilles actives
        wb_principal = load_workbook(principal_filepath)
        ws_principal = wb_principal.active

        wb_source = load_workbook(source_filepath)
        ws_source = wb_source.active

        print(f"Fichier principal chargé: {principal_filepath}")
        print(f"Fichier source chargé: {source_filepath}")

        # Identifier la colonne 'Hôpital' dans le fichier source
        hosp_col_source_idx = find_column(ws_source, ['hôpital', 'hospital', 'hopital', 'HOPITAUX/ CMC'])

        # Déterminer l'hôpital de référence :
        # Prend la première valeur non nulle trouvée dans la colonne hôpital du fichier source,
        # en ignorant l'en-tête.
        reference_hospital_name = None
        for r in range(2, ws_source.max_row + 1):
            cell_value = ws_source.cell(row=r, column=hosp_col_source_idx).value
            if cell_value: # Si la cellule n'est pas vide
                reference_hospital_name = cell_value
                break # Arrête dès que le premier nom d'hôpital est trouvé

        if not reference_hospital_name:
            print(f"⚠️ Aucun nom d'hôpital de référence trouvé dans {source_filepath}. Le fichier sera ignoré.")
            return # Sortie de la fonction si aucun hôpital n'est trouvé

        print(f"Hôpital de référence identifié dans '{source_filepath}': '{reference_hospital_name}'")

        # Récupérer les listes de dates pour l'hôpital de référence
        source_hospital_dates = get_hospital_dates(ws_source, reference_hospital_name)
        principal_hospital_dates = get_hospital_dates(ws_principal, reference_hospital_name)

        if not source_hospital_dates:
            print(f"⚠️ Aucune date trouvée pour l'hôpital '{reference_hospital_name}' dans le fichier source '{source_filepath}'.")
            # On continue quand même au cas où il y aurait des dates dans le principal à comparer (même si missing sera vide)
            # ou on pourrait choisir de retourner ici si c'est une condition bloquante.

        # Identifier les dates manquantes dans le fichier principal
        missing_dates_list = find_missing_dates(source_hospital_dates, principal_hospital_dates)

        if not missing_dates_list:
            print(f"ℹ️ Aucune nouvelle date à importer de '{source_filepath}' pour l'hôpital '{reference_hospital_name}'. "
                  f"Toutes les dates sont déjà présentes dans '{principal_filepath}'.")
            return # Sortie si aucune date manquante

        print(f"Nombre de dates manquantes trouvées pour '{reference_hospital_name}': {len(missing_dates_list)}")

        # Identifier la colonne 'Date' dans le fichier source pour la comparaison des lignes
        date_col_source_idx = find_column(ws_source, ['date', 'DATE', 'Date'])
        # Nombre maximum de colonnes dans le fichier source (pour copier toutes les données de la ligne)
        max_cols_in_source = ws_source.max_column

        # Créer un nouveau classeur et une nouvelle feuille pour les lignes manquantes
        wb_output = Workbook()
        ws_output = wb_output.active
        ws_output.title = f"NouvLignes_{reference_hospital_name[:10]}" # Titre de feuille optionnel

        # Copier l'en-tête
        # Par défaut, l'en-tête du fichier PRINCIPAL est utilisé pour la cohérence.
        # Changez `ws_principal` en `ws_source` si vous préférez l'en-tête du fichier source.
        header_source_sheet = ws_principal # Ou ws_source
        for c in range(1, header_source_sheet.max_column + 1):
            ws_output.cell(row=1, column=c).value = header_source_sheet.cell(row=1, column=c).value

        lines_added_count = 0
        current_output_row = 2 # Commence à écrire à la deuxième ligne du fichier de sortie (après l'en-tête)

        # Parcourir le fichier source pour trouver et copier les lignes manquantes
        for r_src in range(2, ws_source.max_row + 1): # Commence à la ligne 2 pour ignorer l'en-tête
            # Récupérer la date et le nom de l'hôpital de la ligne actuelle du fichier source
            current_row_date_cell = ws_source.cell(row=r_src, column=date_col_source_idx)
            current_row_date = excel_to_datetime(current_row_date_cell)
            current_row_hospital = ws_source.cell(row=r_src, column=hosp_col_source_idx).value

            # Vérifier si la ligne correspond à l'hôpital de référence et à une date manquante
            if current_row_hospital == reference_hospital_name and current_row_date in missing_dates_list:
                # Copier toutes les cellules de cette ligne source vers la feuille de sortie
                for c_src in range(1, max_cols_in_source + 1):
                    source_cell = ws_source.cell(row=r_src, column=c_src)
                    output_cell = ws_output.cell(row=current_output_row, column=c_src)
                    output_cell.value = source_cell.value # Copie la valeur

                    # Tente de préserver le format numérique de la cellule si défini
                    if source_cell.has_style and source_cell.number_format:
                        output_cell.number_format = source_cell.number_format

                current_output_row += 1 # Incrémente l'index de la ligne de sortie
                lines_added_count += 1  # Incrémente le compteur de lignes ajoutées

        # Sauvegarder le nouveau classeur contenant uniquement les lignes manquantes
        if lines_added_count > 0:
            wb_output.save(output_filepath)
            print(f"✅ {lines_added_count} lignes manquantes pour '{reference_hospital_name}' depuis '{source_filepath}' "
                  f"ont été enregistrées dans '{output_filepath}'.")
        else:
            # Ce cas ne devrait pas être atteint si la vérification de `missing_dates_list` est faite avant,
            # mais c'est une sécurité.
            print(f"ℹ️ Aucune ligne n'a été ajoutée pour '{reference_hospital_name}' depuis '{source_filepath}'.")

    except FileNotFoundError:
        print(f"❌ ERREUR CRITIQUE: Fichier non trouvé. Vérifiez les chemins: '{principal_filepath}' ou '{source_filepath}'.")
    except ValueError as ve: # Erreur levée par find_column par exemple
        print(f"❌ ERREUR DE VALEUR lors du traitement de '{source_filepath}': {ve}")
    except Exception as e:
        # Attrape toute autre exception pour éviter l'arrêt brutal du script
        print(f"❌ ERREUR INATTENDUE lors du traitement de '{source_filepath}': {e}")
        import traceback
        traceback.print_exc() # Affiche la trace complète de l'erreur pour le débogage


# --- Point d'Entrée Principal du Script ---

def main():
    """
    Fonction principale qui orchestre la recherche des fichiers Excel sources
    et lance le processus d'extraction des lignes manquantes pour chacun d'eux.
    """
    print("--- Démarrage du script d'extraction des nouvelles lignes Excel ---")

    # 1. Vérifier l'existence du fichier principal
    if not os.path.exists(PRINCIPAL):
        print(f"❌ ERREUR: Le fichier principal '{PRINCIPAL}' est introuvable. Le script ne peut pas continuer.")
        return # Arrêt du script

    # 2. Trouver tous les fichiers .xlsx dans le répertoire courant
    #    Le motif '*.xlsx' signifie "tous les fichiers se terminant par .xlsx"
    all_xlsx_files_in_directory = glob.glob("*.xlsx")

    # 3. Filtrer la liste pour exclure le fichier principal lui-même
    #    et les fichiers de sortie potentiellement déjà générés (pour éviter de les traiter comme sources).
    source_excel_files = [
        f for f in all_xlsx_files_in_directory
        if f != PRINCIPAL and not f.startswith("new_lines_only")
    ]

    # Optionnel: Trier les fichiers sources pour un traitement ordonné (par exemple, alphabétique)
    source_excel_files.sort()

    if not source_excel_files:
        print("ℹ️ Aucun fichier source .xlsx (autre que le principal et les fichiers de sortie) "
              "n'a été trouvé dans le répertoire courant pour traitement.")
        return # Arrêt si aucun fichier source n'est trouvé

    print(f"Fichier principal de référence: '{PRINCIPAL}'")
    print(f"Fichiers sources .xlsx trouvés pour traitement: {source_excel_files}")

    # 4. Boucler sur chaque fichier source trouvé et le traiter
    output_file_index = 1 # Compteur pour nommer les fichiers de sortie de manière unique
    for source_file_path in source_excel_files:
        # Construire le nom du fichier de sortie dynamiquement
        # Exemple: new_lines_only1.xlsx, new_lines_only2.xlsx, ...
        dynamic_output_filename = f"new_lines_only{output_file_index}.xlsx"

        # Appeler la fonction de traitement pour le fichier source actuel
        process_files(PRINCIPAL, source_file_path, dynamic_output_filename)

        output_file_index += 1 # Incrémenter pour le prochain fichier de sortie

    print("\n--- Traitement de tous les fichiers terminé. ---")


if __name__ == '__main__':
    # Cette condition vérifie si le script est exécuté directement (et non importé comme module)
    main()