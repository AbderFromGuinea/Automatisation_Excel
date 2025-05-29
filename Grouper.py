#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import glob
import os
from datetime import datetime  # Non utilisé activement mais bon à avoir pour les dates
import traceback  # Pour des traces d'erreur plus complètes

# Format de date courte souhaité pour Excel (ex: 20/05/2025)
SHORT_DATE_FORMAT = 'dd/mm/yyyy'


def get_or_create_groupe_sheet():
    """
    Charge le fichier 'groupe.xlsx' s'il existe, sinon le crée.
    Retourne le classeur (workbook) et la feuille active (worksheet).
    """
    filename = 'groupe.xlsx'
    try:
        wb_dest = openpyxl.load_workbook(filename)
        print(f"INFO: Fichier destination '{filename}' existant chargé.")
    except FileNotFoundError:
        wb_dest = openpyxl.Workbook()
        print(f"INFO: Nouveau fichier destination '{filename}' créé.")

    ws_dest = wb_dest.active
    if ws_dest.title == "Sheet":
        ws_dest.title = "DonneesGroupees"
    return wb_dest, ws_dest


def find_first_empty_row(ws):
    """
    Trouve l'index de la première ligne vide dans une feuille de calcul.
    """
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        return 1
    return ws.max_row + 1


def process_all_excel_files():
    """
    Fonction principale pour traiter tous les fichiers Excel éligibles
    et regrouper leurs données dans 'groupe.xlsx'.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)
    print(f"INFO: Répertoire de travail actuel : {os.getcwd()}")

    all_xlsx_files = glob.glob('*.xlsx')
    print(f"DEBUG: Fichiers .xlsx trouvés par glob : {all_xlsx_files}")

    # Exclure 'recette.xlsx' (insensible à la casse)
    target_files = []
    for f_name in all_xlsx_files:
        if f_name.lower() == 'Ventes1.xlsx':
            print(f"DEBUG: '{f_name}' est explicitement exclu (correspond à 'recette.xlsx').")
        else:
            target_files.append(f_name)

    print(f"INFO: Fichiers .xlsx qui seront traités (après filtrage) : {target_files}")

    if not target_files:
        print("INFO: Aucun fichier .xlsx à traiter (autre que recette.xlsx) n'a été trouvé.")
        return

    wb_groupe, ws_groupe = get_or_create_groupe_sheet()
    next_row_in_groupe = find_first_empty_row(ws_groupe)
    source_column_indices = [2, 3, 4, 7, 8, 9]  # B, C, D, G, H, I
    files_processed_count = 0
    total_rows_added_to_groupe = 0

    for filename in target_files:
        # Sécurité supplémentaire: ignorer 'recette.xlsx' au cas où le filtre initial aurait un problème
        # pour un nom de fichier très particulier.
        if filename.lower() == 'ventes1.xlsx':
            print(f"  AVERTISSEMENT CRITIQUE: '{filename}' a contourné le filtre initial mais est bloqué ici. Ignoré.")
            continue

        print(f"INFO: Traitement du fichier source : {filename}...")

        try:
            wb_source = openpyxl.load_workbook(filename)  # data_only=False est généralement mieux pour les dates
            ws_source = wb_source.active

            if ws_source.max_row == 0 or \
                    (ws_source.max_row == 1 and ws_source.cell(row=1,
                                                               column=1).value is None and ws_source.max_column == 1):
                print(f"  INFO: Le fichier '{filename}' semble vide. Ignoré.")
                continue

            print(f"  INFO: '{filename}' contient {ws_source.max_row} lignes. Copie des données...")
            rows_copied_from_this_file = 0

            for row_num in range(1, ws_source.max_row + 1):
                if row_num % 500 == 0:  # Imprimer une progression pour les gros fichiers
                    print(f"    DEBUG: Traitement de la ligne {row_num}/{ws_source.max_row} de '{filename}'...")

                data_from_current_row = []
                has_any_data_in_selected_cols = False
                for col_idx_source in source_column_indices:
                    cell_value = None
                    if col_idx_source <= ws_source.max_column:
                        cell_value = ws_source.cell(row=row_num, column=col_idx_source).value
                    data_from_current_row.append(cell_value)
                    if cell_value is not None:
                        has_any_data_in_selected_cols = True

                if not has_any_data_in_selected_cols and all(v is None for v in data_from_current_row):
                    continue  # Sauter les lignes entièrement vides pour les colonnes sélectionnées

                for col_idx_dest, value_to_write in enumerate(data_from_current_row, start=1):
                    dest_cell = ws_groupe.cell(row=next_row_in_groupe, column=col_idx_dest)
                    dest_cell.value = value_to_write
                    if col_idx_dest == 1:  # Colonne B (date)
                        dest_cell.number_format = SHORT_DATE_FORMAT

                next_row_in_groupe += 1
                rows_copied_from_this_file += 1

            if rows_copied_from_this_file > 0:
                print(f"  INFO: {rows_copied_from_this_file} lignes de '{filename}' ont été ajoutées à 'groupe.xlsx'.")
                total_rows_added_to_groupe += rows_copied_from_this_file
            else:
                print(f"  INFO: Aucune ligne pertinente n'a été copiée de '{filename}'.")

            files_processed_count += 1

        except Exception as e:
            print(f"  ERREUR: Problème lors du traitement du fichier '{filename}': {e}")
            traceback.print_exc()  # Imprime la trace complète de l'erreur

    print(f"\nINFO: Fin de la boucle de traitement des fichiers. {files_processed_count} fichier(s) ont été traités.")
    print(f"INFO: Total de {total_rows_added_to_groupe} lignes ajoutées à la feuille de destination en mémoire.")
    print("INFO: Tentative de sauvegarde de 'groupe.xlsx'...")

    try:
        wb_groupe.save('groupe.xlsx')
        print(f"\nSUCCÈS: Opération terminée. Toutes les données ont été compilées dans 'groupe.xlsx'.")
    except Exception as e:
        print(f"  ERREUR DÉFINITIVE lors de la sauvegarde de 'groupe.xlsx': {e}")
        traceback.print_exc()  # Imprime la trace complète de l'erreur


if __name__ == "__main__":
    process_all_excel_files()

